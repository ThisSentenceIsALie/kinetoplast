'''
Program name: dissolusion_plot.py
Author: Nathaniel Morrison
Date created: 05/30/2020
Date last modified: 01/15/2023
Python Version: 3.7.2
'''
'''
Generates a plot of bf_search data
'''
import openpyxl
import matplotlib.pyplot as plt

#Function to load bf_search output datafile
def excel_loader(io_name):
    #Load the workbook
    wb=openpyxl.load_workbook(io_name+'.xlsx')
    #Return the workbook and worksheet
    return wb, wb[wb.sheetnames[0]]

#Function to gather the x-y coordinate data from the datafile
def data_retriever(ws,flag_bound,thresh):
    #Get total number of rows. For whatever reason the indexing starts at 1 in this module, not 0
    tot_rows=ws.max_row
    #Row 1 is header. Start at row 2
    cur_row=2
    x_coords=list()
    y_coords=list()
    bound_break_x1=''
    bound_break_y1=''
    bound_break_x2=''
    bound_break_y2=''
    #Run through each row
    while cur_row<=tot_rows:
        #The x-coordinate (number of dissolutions) is the same for all points in the row
        x_coord=int(ws.cell(cur_row,1).value)
        #Colummn 1 is the number of dissolutions. Starting with column 2, each entry is a pair: (component size, number of components with this size)
        cur_col=2
        #Get the pair as a string
        cur_pair=ws.cell(cur_row,cur_col).value
        #An empty cell returns None type; go until we run into that
        while cur_pair!=None and cur_pair!='Boundary Breaks Here' and cur_pair!='Boundary Splits in Two Here':
            #Convert string to a tuple. The first element is the component size (which we want) and the second is the number of componenets of that size (which we don't)
            try:
                cur_pair=eval(cur_pair)
            except:
                print(type(cur_pair))
                int('a')
            #Omit components with size < the threshold size
            if cur_pair[0]>=thresh:
                y_coords.append(cur_pair[0])
                x_coords.append(x_coord)
            cur_col+=1
            cur_pair=ws.cell(cur_row,cur_col).value
        #If the user wants to flag where the boundary first breaks apart
        if flag_bound:
            #Find the row in the datafile with the phrase 'Boundary Breaks Here' printed after the list of component sizes
            if cur_pair=='Boundary Breaks Here':
                bound_break_x1=list()
                bound_break_y1=list()
                #Take note of the x/y values at this step.
                for i in range(len(x_coords)):
                    if x_coords[i]==x_coord:
                        bound_break_y1.append(y_coords[i])
                        bound_break_x1.append(x_coord)
            #Do same for the row with the phrase 'Boundary Splits in Two Here'
            elif cur_pair=='Boundary Splits in Two Here':
                bound_break_x2=list()
                bound_break_y2=list()
                for i in range(len(x_coords)):
                    if x_coords[i]==x_coord:
                        bound_break_y2.append(y_coords[i])
                        bound_break_x2.append(x_coord)
        cur_row+=1
    #Return the x and y coordinate tuples
    return tuple(x_coords),tuple(y_coords),bound_break_x1,bound_break_y1,bound_break_x2,bound_break_y2

#Function do the actual plotting
def plotter(x_coords, y_coords,io_name,bound_break_x1,bound_break_y1,bound_break_x2,bound_break_y2,flag_bound):
    #Create the figure object
    fig = plt.figure()
    #Generate a scatter plot with circle markers. s is the area of the markers in arbitrary units; it defaults to 20
    plt.scatter(x_coords,y_coords,marker='o',s=0.5)
    #If the user wants to flag the boundary dissintigration
    if flag_bound:
        #Mark the points at the step when the boundary breaks in red. And make them a tad bigger
        plt.scatter(bound_break_x1,bound_break_y1,marker='o',s=1.5,c='r')
        #Make the points at the step when the boundary lyses into two sperate pieces green. And make them a tad bigger
        plt.scatter(bound_break_x2,bound_break_y2,marker='o',s=1.5,c='lime')
    #Generate the plot title
    fig.suptitle(io_name, fontsize=20)
    #Generate the axis labels
    plt.xlabel('Dissolutions', fontsize=18)
    plt.ylabel('Nodes in Component(s)', fontsize=16)
    #Save the figure with 300 dots per inch
    fig.savefig(io_name+'_plot.png',dpi=300)
    plt.clf()
    return

#Function to allow this program to act as an externally-called module rather than a script
def module_main(pic_io_name,data_io_name,flag_bound=False,thresh=10):
    #Grab the input datafile
    wb,ws=excel_loader(data_io_name)
    #Extract point coordinates from the datafile
    x_coords, y_coords,bound_break_x1,bound_break_y1,bound_break_x2,bound_break_y2=data_retriever(ws,flag_bound,thresh)
    wb.close()
    #Generate the plot
    plotter(x_coords,y_coords,pic_io_name,bound_break_x1,bound_break_y1,bound_break_x2,bound_break_y2,flag_bound)
    return

#Function to conduct this orchestra of chaos
if __name__=='__main__':
    flag_bound=False
    #Define the name of the both the input and output files
    io_name="20200607_RKP_NM_hexagonalGrid_10"
    #Grab the input datafile
    wb,ws=excel_loader(io_name)
    #Extract point coordinates from the datafile
    x_coords, y_coords,bound_break_x1,bound_break_y1,bound_break_x2,bound_break_y2=data_retriever(ws,flag_bound)
    #Generate the plot
    plotter(x_coords,y_coords,io_name,bound_break_x1,bound_break_y1,bound_break_x2,bound_break_y2,flag_bound)
#main()
