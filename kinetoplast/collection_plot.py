'''
File name: new_search_manager.py
Author: Nathaniel Morrison
Date created: 01/10/2023
Date last modified: 01/16/2023
Python Version: 3.11.1
'''
'''
Like collection_plot, but better in every way. Switched from xlrd to openpyxl, and now compatible with new file heirarchy
'''

import openpyxl
import os
from matplotlib.collections import LineCollection
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D
import numpy as np

class collectionPlot:
    def __init__(self,folder,autogenFolder=False):
        if folder[-1]!='\\':
            folder+='\\'
        self.folder=folder
        self.autogenFolder=autogenFolder
        if autogenFolder:
            self.infolder=folder+'data\\'
            self.outfolder=folder+'pics\\'
        else:
            self.infolder=folder
            self.outfolder=folder
        self.all_dat=self._load_data()
        return

    def _load_data(self):
        all_dat={}
        all_x={}
        all_y={}
        fileLis=os.listdir(self.infolder)
        for f in fileLis:
            if self._isExcel(f):
                f2=f.rstrip('.xlsx')
                f_lis=f2.split('sat')
                satstr=''
                for a in f_lis[-1]:
                    if a.isnumeric():
                        satstr+=a
                sat=int(satstr)
                all_dat[f]=['','',sat]
                all_dat[f][0],all_dat[f][1]=self.read_data(f)
        return all_dat

    def _isExcel(self,f):
        fLis=f.split('.')
        ext=fLis[1]
        if ext=='xlsx' or ext=='xlsm' or ext=='xlts' or ext=='xltm' or ext=='ods':
            return True
        else:
            return False
    #Function to load bf_search output datafile
    def excel_loader(self,io_name):
        #Load the workbook
        wb=openpyxl.load_workbook(io_name)
        #Return the workbook and worksheet
        return wb, wb[wb.sheetnames[0]]
        
    def read_data(self,f):
        wb,ws=self.excel_loader(self.infolder+f)
        #Get total number of rows. For whatever reason the indexing starts at 1 in this module, not 0
        tot_rows=ws.max_row
        #Row 1 is header. Start at row 2
        cur_row=2
        x_coords=list()
        y_coords=list()
        while cur_row<=tot_rows:
            #The x-coordinate (number of dissolutions) is the same for all points in the row
            x_coord=int(ws.cell(cur_row,1).value)
            x_coords.append(x_coord)
            #Colummn 1 is the number of dissolutions. Starting with column 2, each entry is a pair: (component size, number of components with this size)
            cur_col=2
            #Get the pair as a string
            cur_y=ws.cell(cur_row,cur_col).value
            y_coords_cur=[]
            #An empty cell returns None type; go until we run into that
            while cur_y!=None :
                y_coords_cur.append(cur_y)
                cur_col+=1
                cur_y=ws.cell(cur_row,cur_col).value
            y_coords.append(tuple(y_coords_cur))
            del y_coords_cur[:]
            cur_row+=1
        wb.close()
        return tuple(x_coords),tuple(y_coords)

#Code courtesy of Brendan Artley
    def hex_to_RGB(self,hex_str):
        """ #FFFFFF -> [255,255,255]"""
        #Pass 16 to the integer function for change of base
        return [int(hex_str[i:i+2], 16) for i in range(1,6,2)]

    def get_color_gradient(self,c1, c2, n):
        """
        Given two hex colors, returns a color gradient
        with n colors.
        """
        assert n > 1
        c1_rgb = np.array(self.hex_to_RGB(c1))/255
        c2_rgb = np.array(self.hex_to_RGB(c2))/255
        mix_pcts = [x/(n-1) for x in range(n)]
        rgb_colors = [((1-mix)*c1_rgb + (mix*c2_rgb)) for mix in mix_pcts]
        return ["#" + "".join([format(int(round(val*255)), "02x") for val in item]) for item in rgb_colors]
#End of code courtesy of Brendan Artley

    def plot_component_collection(self,component_number):
        #Create the figure
        fig = plt.figure()
        ax = fig.add_subplot(111, projection='3d')
        dx_master=list()
        dy_master=list()
        dx_all=list()
        dy_all=list()
        dz_all=list()
        c=self.get_color_gradient('DC1960','1E17D8',len(list(self.all_dat.keys())))
        j=0
        for key in self.all_dat.keys():
            cur_dat=self.all_dat[key]
            dx_all.append(cur_dat[0][0])
            dz_all.append(cur_dat[2])
            for i in range(len(cur_dat[0])):
                dx_master.append(cur_dat[0][i])
                dy_master.append(cur_dat[1][i][component_number-1])
                dy_all.append(cur_dat[1][i][component_number-1])
            dx_all.append(cur_dat[0][i])
            ax.plot(dx_master,dy_master,zs=cur_dat[2],zdir='x',marker='o',ms=0.5,color=c[j])
            del dx_master[:]
            del dy_master[:]
            j+=1
        i=0
        #Set the ranges of the plot
        ax.set_xlim(min(dz_all),max(dz_all))
        ax.set_ylim(min(dx_all), max(dx_all))
        ax.set_zlim(0,max(dy_all)*1.15)
        #Add plot/axis titles
        ax.set_xlabel('Nodes per Boundary Stack')
        ax.set_ylabel('Dissolutions')
        compNum=str(component_number)
        if compNum=='1':
            compNum=''
        elif compNum[-1]=='1':
            compNum+='st '
        elif compNum[-1]=='2':
            compNum+='nd '
        elif compNum[-1]=='3':
            compNum+='rd '
        else:
            compNum+='th'
        ax.set_zlabel('Nodes in '+compNum+'Largest Component')
        #ax.set_title("Dependence of Component Dissolution on Boundary Saturation")
        #Show the plot
        plt.show()

if __name__=='__main__':
    cpt=collectionPlot('C:\\Users\\natha\\OneDrive\\Desktop\\CSULB\\kinetoplast\\to_aggregate')
    cpt.plot_component_collection(2)
