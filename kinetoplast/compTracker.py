'''
Module name: compTracker.py
Author: Nathaniel Morrison
Date created: 01/15/2023
Date last modified: 01/15/2023
Python Version: 3.11.1
'''
import openpyxl
from matplotlib import pyplot as plt
import os
import numpy as np

class compTracker:
    def __init__(self,folder,autogenFolder=False,thresh=0):
        if folder[-1]!='\\':
            folder+='\\'
        self.folder=folder
        self.autogenFolder=autogenFolder
        if autogenFolder:
            self.infolder=folder+'data\\'
            self.outfolder=folder+'pics\\'
        else:
            self.infolder=folder
            self.outfolder=folder
        self.thresh=thresh
        self.max_x=0
        self.min_x=np.inf
        self.max_num_y=0
        self.num_f=0
        self.all_dat=self._load_data()
        self.averager()
        return


    def _load_data(self):
        all_dat={}
        all_x={}
        all_y={}
        fileLis=os.listdir(self.infolder)
        for f in fileLis:
            if self._isExcel(f):
                self.num_f+=1
                all_dat[f]=['','']
                all_dat[f][0],all_dat[f][1]=self.read_data(f)
        return all_dat

    def _isExcel(self,f):
        fLis=f.split('.')
        ext=fLis[1]
        if ext=='xlsx' or ext=='xlsm' or ext=='xlts' or ext=='xltm' or ext=='ods':
            return True
        else:
            return False
        
    def read_data(self,f):
        wb,ws=self.excel_loader(self.infolder+f)
        #Get total number of rows. For whatever reason the indexing starts at 1 in this module, not 0
        tot_rows=ws.max_row
        #Row 1 is header. Start at row 2
        cur_row=2
        x_coords=list()
        y_coords=list()
        while cur_row<=tot_rows:
            #The x-coordinate (number of dissolutions) is the same for all points in the row
            x_coord=int(ws.cell(cur_row,1).value)
            x_coords.append(x_coord)
            #Colummn 1 is the number of dissolutions. Starting with column 2, each entry is a pair: (component size, number of components with this size)
            cur_col=2
            #Get the pair as a string
            cur_pair=ws.cell(cur_row,cur_col).value
            y_coords_cur=[]
            #An empty cell returns None type; go until we run into that
            while cur_pair!=None and cur_pair!='Boundary Breaks Here' and cur_pair!='Boundary Splits in Two Here':
                #Convert string to a tuple. The first element is the component size (which we want) and the second is the number of componenets of that size (which we don't)
                cur_pair=eval(cur_pair)
                #Omit components with size < the threshold size
                if cur_pair[0]>=self.thresh:
                    y_coords_cur.append(cur_pair[0])
                cur_col+=1
                cur_pair=ws.cell(cur_row,cur_col).value
            y_coords.append(tuple(y_coords_cur))
            del y_coords_cur[:]
            if cur_col-2>self.max_num_y:
                self.max_num_y=cur_col-2
            cur_row+=1
        if cur_row-2>self.max_x:
            self.max_x=cur_row-2
        if cur_row-2<self.min_x:
            self.min_x=cur_row-2
        wb.close()
        return tuple(x_coords),tuple(y_coords)

    #Only data up to the run with the smallest number of dissolution steps will be used
    def averager(self):
        sum_y=[]
        for i in range(self.min_x):
            sum_y.append([0,]*self.max_num_y)
        self.avg_x=[]
        getx=True
        for key in self.all_dat.keys():
            cur_dat=self.all_dat[key]
            for i in range(self.min_x):
                if getx:
                    self.avg_x.append(cur_dat[0][i])
                cur_y=cur_dat[1][i]
                for j in range(len(cur_y)):
                    sum_y[i][j]+=cur_y[j]
            getx=False
        for k in range(len(sum_y)):
            sum_y[k]=tuple(np.array(sum_y[k])/self.num_f)
        self.avg_y=tuple(sum_y)
        return
        
    #Function to load bf_search output datafile
    def excel_loader(self,io_name):
        #Load the workbook
        wb=openpyxl.load_workbook(io_name)
        #Return the workbook and worksheet
        return wb, wb[wb.sheetnames[0]]

    def plotLargestComp(self):
        ydat=list(k[0] for k in self.avg_y)
        #Create the figure object
        fig = plt.figure()
        #Generate a scatter plot with circle markers. s is the area of the markers in arbitrary units; it defaults to 20
        plt.scatter(self.avg_x,ydat,marker='o',s=0.5)
        #Generate the axis labels
        plt.xlabel('Dissolutions', fontsize=18)
        plt.ylabel('Average Nodes in Largest Component', fontsize=16)
        #Save the figure with 300 dots per inch
        fig.savefig(self.outfolder+'avgLargestCompTracker.png',dpi=300)
        plt.clf()
        return

    def plotSecondLargestComp(self):
        ydat=list(k[1] for k in self.avg_y)
        #Create the figure object
        fig = plt.figure()
        #Generate a scatter plot with circle markers. s is the area of the markers in arbitrary units; it defaults to 20
        plt.scatter(self.avg_x,ydat,marker='o',s=0.5)
        #Generate the axis labels
        plt.xlabel('Dissolutions', fontsize=18)
        plt.ylabel('Average Nodes in 2nd Largest Component', fontsize=16)
        #Save the figure with 300 dots per inch
        fig.savefig(self.outfolder+'avg2ndLargestCompTracker.png',dpi=300)
        plt.clf()
        return

    def export_data(self,outfolder='',extraName=''):
        if outfolder=='':
            outfolder=self.infolder
        if outfolder[-1]!='\\':
            outfolder+='\\'
        wb=openpyxl.Workbook()
        ws=wb.active
        ws['A1']='Dissolutions'
        for k in range(2,len(self.avg_y[0])+2):
            ws.cell(row=1,column=k).value='Average Size of '+str(k)+' Component'
        for i in range(len(self.avg_x)):
            x=self.avg_x[i]
            y=self.avg_y[i]
            ws.cell(row=i+2, column=1).value=x
            for k in range(len(y)):
                ws.cell(row=i+2,column=k+2).value=y[k]
        wb.save(outfolder+'average_component_sizes'+extraName+'.xlsx')
        wb.close()
        return
            
        









    

