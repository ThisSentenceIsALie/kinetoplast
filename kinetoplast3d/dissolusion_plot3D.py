'''
Program name: dissolusion_plot.py
Author: Nathaniel Morrison
Date created: 06/29/2020
Date last modified: 01/16/2023
Python Version: 3.11.1
'''
'''
Generates a plot of bf_search data
'''
import openpyxl
import matplotlib.pyplot as plt

#Function to load bf_search output datafile
def excel_loader(io_name):
    #Load the workbook
    wb=openpyxl.load_workbook(io_name+'.xlsx')
    #Return the workbook and worksheet
    return wb, wb[wb.sheetnames[0]]

):
    #Get total number of rows. For whatever reason the indexing starts at 1 in this module, not 0
    tot_rows=ws.max_row
    #Row 1 is header. Start at row 2
    cur_row=2
    x_coords=list()
    y_coords=list()
    bound_break_x1=''
    bound_break_y1=''
    bound_break_x2=''
    bound_break_y2=''
    #Run through each row
    while cur_row<=tot_rows:
        #The x-coordinate (number of dissolutions) is the same for all points in the row
        x_coord=int(ws.cell(cur_row,1).value)
        #Colummn 1 is the number of dissolutions. Starting with column 2, each entry is a pair: (component size, number of components with this size)
        cur_col=2
        #Get the pair as a string
        cur_pair=ws.cell(cur_row,cur_col).value
        #An empty cell returns None type; go until we run into that
        while cur_pair!=None:
            #Convert string to a tuple. The first element is the component size (which we want) and the second is the number of componenets of that size (which we don't)
            try:
                cur_pair=eval(cur_pair)
            except:
                print(type(cur_pair))
                int('a')
            #Omit components with size < the threshold size
            if cur_pair[0]>=thresh:
                y_coords.append(cur_pair[0])
                x_coords.append(x_coord)
            cur_col+=1
            cur_pair=ws.cell(cur_row,cur_col).value
        cur_row+=1
    #Return the x and y coordinate tuples
    return tuple(x_coords),tuple(y_coords)

#Function do the actual plotting
def plotter(x_coords, y_coords,io_name):
    #Create the figure object
    fig = plt.figure()
    #Generate a scatter plot with circle markers. s is the area of the markers in arbitrary units; it defaults to 20
    plt.scatter(x_coords,y_coords,marker='o',s=0.5)
    #Generate the plot title
    fig.suptitle(io_name, fontsize=20)
    #Generate the axis labels
    plt.xlabel('Dissolutions', fontsize=18)
    plt.ylabel('Nodes in Component(s)', fontsize=16)
    #Save the figure with 300 dots per inch
    fig.savefig(io_name+'_plot.png',dpi=300)
    return

#Function to allow this program to act as an externally-called module rather than a script
def module_main(pic_io_name,data_io_name,thresh=0):
    #Grab the input datafile
    wb,ws=excel_loader(data_io_name)
    #Extract point coordinates from the datafile
    x_coords, y_coords=data_retriever(ws,thresh)
    wb.close()
    #Generate the plot
    plotter(x_coords,y_coords,pic_io_name)
    return
